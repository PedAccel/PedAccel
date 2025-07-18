'''
Extracts Nurse SBS Scores from CCDA Data
|_ To be run on JHH SAFE Desktop.
'''

# Import Modules
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
import os
import re
os.chdir(r'S:\Sedation_monitoring')

def load_ccda_sbs():
    '''
    Load the CCDA SBS Data for all Patients

    Note: These xlsx files have NOT been edited for interference.

    Warning: Do not push this file with MRN information to GITHUB.
    '''
    patient_info_add = 'Full_Patient_List_CCDA.xlsx'
    patient_info_excel = load_workbook(patient_info_add)
    patient_info = patient_info_excel['Sheet1']

    print('Patient List Loaded')

    for cell_a, cell_b in zip(patient_info['A'][1:], patient_info['B'][1:]):
        patient_num = cell_a.value
        patient_mrn = cell_b.value

        print(f'Processing: Patient {patient_num}')
        
        # Specify the path to the Excel file
        file_path = 'CCDA_6771_Extract_03042024.xlsx'

        # Load the Excel workbook
        workbook = load_workbook(file_path)
        worksheet = workbook['ABCDEF Bundle']

        # Make Blank 2D Array
        sbs_scores = []

        for cell_c, cell_i, cell_k in zip(worksheet['C'][1:], worksheet['I'][1:], worksheet['K'][1:]):
            # Access the value of the cells in columns C and I
            value_c = cell_c.value
            value_i = cell_i.value
            value_k = cell_k.value
    
            # Check if the value in column C matches the desired value
            if value_c == patient_mrn:
                if value_i[0] == '0' or value_i[1] in ['1', '2', '3']:
                    time_object = value_k
                    excel_dt = time_object.strftime('%Y-%m-%d %H:%M:%S')

                    cleaned_value_i = re.sub(r'[^\d+-]', '', value_i)
                    sbs_scores.append([excel_dt, cleaned_value_i])
            
        print(sbs_scores)
        sbs_scores.sort(key=lambda x: x[0])

        newscores = Workbook()
        newscores_worksheet = newscores.active

        newscores_worksheet['A1'] = "Time_uniform"
        newscores_worksheet['A1'].font = Font(bold=True)

        newscores_worksheet['B1'] = "Time"
        newscores_worksheet['B1'].font = Font(bold=True)

        newscores_worksheet['C1'] = "SBS"
        newscores_worksheet['C1'].font = Font(bold=True)

        row_index = 2
        for timestamp, sbs_score in enumerate(sbs_scores, start=2):  # Starting from the second row after titles
            newscores_worksheet.cell(row_index, 3, value=float(sbs_score[1]))
            newscores_worksheet.cell(row_index, 2, value=sbs_score[0])

            time_uniform_formula = f'=TEXT(B{row_index}, "M/dd/yyyy hh:mm::ss AM/PM")'
            newscores_worksheet.cell(row_index, 1).value = time_uniform_formula
            row_index += 1

        # Save the workbook
        filename = f'Patient{patient_num}_SBS_Scores.xlsx'
        patient_dir = r"S:\Sedation_monitoring\Nurse_SBS_CCDA"
        savefile = os.path.join(patient_dir, filename)
        newscores.save(savefile)

        print(f"Patient {patient_num} data recorded")
        workbook.close()

def load_sbs_goal():
    
    patient_info_add = 'Full_Patient_List_CCDA.xlsx'
    patient_info_excel = load_workbook(patient_info_add)
    patient_info = patient_info_excel['Sheet2']

    print('Patient List Loaded')

    for cell_a, cell_b, cell_c, cell_d in zip(patient_info['A'][1:], patient_info['B'][1:], patient_info['C'][1:], patient_info['D'][1:]):
        patient_num = cell_a.value
        patient_mrn = cell_b.value
        patient_newid = cell_c.value
        patient_enc_newid = cell_d.value

        print(f'Processing: Patient {patient_num}')
        
        # Specify the path to the Excel file
        file_path = r'S:\Sedation_monitoring\CCDA_6771_extract_IDENTIFIABLE_HIPPA\ccda6771_02192025.xlsx'

        # Load the Excel workbook
        workbook = load_workbook(file_path)
        worksheet = workbook['ABCDEF Bundle']

        # Make Blank 2D Array
        sbs_scores = []

        sbs_goals = []

        for cell_c, cell_e, cell_f, cell_g in zip(worksheet['C'][1:], worksheet['E'][1:], worksheet['F'][1:],worksheet['G'][1:]):
            # Access the value of the cells in columns C and I
            value_c = cell_c.value
            value_e = cell_e.value
            value_f = cell_f.value
            value_g = cell_g.value
    
            # Check if the value in column C matches the desired value
            if value_c == patient_newid:
                if value_e[0] == '0' or value_e[1] in ['1', '2', '3']:
                    time_object = value_g
                    excel_dt = time_object.strftime('%Y-%m-%d %H:%M:%S')

                    cleaned_value_i = re.sub(r'[^\d+-]', '', value_e)
                    sbs_scores.append([excel_dt, cleaned_value_i])

                if value_e == "Target SBS Level":
                    if value_f[0] == '0' or value_f[1] in ['1', '2', '3']:
                        time_object = value_g
                        excel_dt = time_object.strftime('%Y-%m-%d %H:%M:%S')

                        cleaned_value_j = re.sub(r'[^\d+-]', '', value_f)
                        sbs_goals.append([excel_dt, cleaned_value_j])
            
        print(sbs_scores)
        print(sbs_goals)
        sbs_scores.sort(key=lambda x: x[0])
        sbs_goals.sort(key=lambda x: x[0])

        newscores = Workbook()
        newscores_worksheet = newscores.active

        newscores_worksheet['A1'] = "Time_uniform"
        newscores_worksheet['A1'].font = Font(bold=True)

        newscores_worksheet['B1'] = "Time"
        newscores_worksheet['B1'].font = Font(bold=True)

        newscores_worksheet['C1'] = "SBS"
        newscores_worksheet['C1'].font = Font(bold=True)

        newscores_worksheet['D1'] = "SBS_Goal"
        newscores_worksheet['D1'].font = Font(bold=True)

        row_index = 2
        for timestamp, sbs_score in enumerate(sbs_scores, start=2):  # Starting from the second row after titles
            newscores_worksheet.cell(row_index, 3, value=float(sbs_score[1]))
            newscores_worksheet.cell(row_index, 2, value=sbs_score[0])

            time_uniform_formula = f'=TEXT(B{row_index}, "M/dd/yyyy hh:mm::ss AM/PM")'
            newscores_worksheet.cell(row_index, 1).value = time_uniform_formula
            row_index += 1

        # Add SBS Goals to the spreadsheet after SBS Scores
        for sbs_goal in sbs_goals:  # Iterate through SBS goals
            newscores_worksheet.cell(row_index, 1, value=f'=TEXT("{sbs_goal[0]}", "M/dd/yyyy hh:mm::ss AM/PM")')
            newscores_worksheet.cell(row_index, 2, value=sbs_goal[0])
            newscores_worksheet.cell(row_index, 4, value=float(sbs_goal[1]))
            row_index += 1

        # Save the workbook
        filename = f'Patient{patient_num}_SBS_Scores.xlsx'
        patient_dir = r"S:\Sedation_monitoring\Nurse_SBS_CCDA"
        savefile = os.path.join(patient_dir, filename)
        newscores.save(savefile)

        print(f"Patient {patient_num} data recorded")
        workbook.close()

if __name__ == '__main__':
    # load_ccda_sbs()

    load_sbs_goal()